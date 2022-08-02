from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Release Year','IMDB Rating'])

source = requests.get("https://www.imdb.com/chart/top")
source.raise_for_status()
soup = BeautifulSoup(source.text, 'lxml')
movies= soup.find('tbody', class_="lister-list").find_all('tr')
for movie in movies:
    name = movie.find('td', class_="titleColumn").a.text
    rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
    year = movie.find('td', class_="titleColumn").span.text.strip('()')
    rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
    print(rank , name, year, rating)
    sheet.append([rank, name, year, rating])

excel.save('IMDB Top Rated Movies.xlsx')

